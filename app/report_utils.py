from io import BytesIO

from flask import (
    request,
    send_file
)
from sqlalchemy import func, extract

from datetime import date

from openpyxl import Workbook

from openpyxl.styles import (
    Font,
    PatternFill
)

from app.extensions import db

from reportlab.lib.pagesizes import landscape, A4

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)

from reportlab.lib import colors

from reportlab.lib.styles import (
    getSampleStyleSheet
)

from reportlab.lib.enums import (
    TA_CENTER
)

from reportlab.lib.units import inch

from sqlalchemy.orm import joinedload

from .models import (
    Inspection,
    Deviation,
    CableBatch,
    Customer,
    CableType,
    ProductionLine,
    CAPA,
    QualityMetric,
    User,
    AuditLog,
    Company

)

def get_filtered_inspection_query(company_id):

    query = Inspection.query.options(

        joinedload(
            Inspection.batch
        ).joinedload(
            CableBatch.customer
        ),

        joinedload(
            Inspection.batch
        ).joinedload(
            CableBatch.cable_type
        ),

        joinedload(
            Inspection.batch
        ).joinedload(
            CableBatch.production_line
        )

    ).filter(

        Inspection.company_id == company_id

    )

    date_from = request.args.get(
        "date_from"
    )

    date_to = request.args.get(
        "date_to"
    )

    inspector = request.args.get(
        "inspector"
    )

    status = request.args.get(
        "status"
    )

    customer = request.args.get(
        "customer"
    )

    cable_type = request.args.get(
        "cable_type"
    )

    production_line = request.args.get(
        "production_line"
    )

    batch_number = request.args.get(
        "batch_number"
    )

    if date_from:

        query = query.filter(

            Inspection.inspection_date >= date_from

        )

    if date_to:

        query = query.filter(

            Inspection.inspection_date <= date_to

        )

    if inspector:

        query = query.filter(

            Inspection.inspector.ilike(

                f"%{inspector}%"

            )

        )

    if status:

        query = query.filter(

            Inspection.overall_result == status

        )

    if customer:

        query = query.join(

            Inspection.batch

        ).join(

            CableBatch.customer

        ).filter(

            Customer.company_name.ilike(

                f"%{customer}%"

            )

        )

    if cable_type:

        query = query.join(

            Inspection.batch

        ).join(

            CableBatch.cable_type

        ).filter(

            CableType.name.ilike(

                f"%{cable_type}%"

            )

        )

    if production_line:

        query = query.join(

            Inspection.batch

        ).join(

            CableBatch.production_line

        ).filter(

            ProductionLine.line_name.ilike(

                f"%{production_line}%"

            )

        )

    if batch_number:

        query = query.join(

            Inspection.batch

        ).filter(

            CableBatch.batch_number.ilike(

                f"%{batch_number}%"

            )

        )

    return query

def apply_inspection_sort(query):

    sort = request.args.get(

        "sort",

        "inspection_date"

    )

    direction = request.args.get(

        "direction",

        "desc"

    )

    sort_columns = {

        "inspection_number": Inspection.inspection_number,

        "inspection_date": Inspection.inspection_date,

        "inspector": Inspection.inspector,

        "result": Inspection.overall_result

    }

    sort_column = sort_columns.get(

        sort,

        Inspection.inspection_date

    )

    if direction == "asc":

        query = query.order_by(

            sort_column.asc()

        )

    else:

        query = query.order_by(

            sort_column.desc()

        )

    return query

def paginate_query(

        query,

        per_page=5

):

    page = request.args.get(

        "page",

        1,

        type=int

    )

    pagination = query.paginate(

        page=page,

        per_page=per_page,

        error_out=False

    )

    return pagination


def get_inspection_statistics(query):

    total = query.count()

    passed = query.filter(

        Inspection.overall_result == "Pass"

    ).count()

    failed = query.filter(

        Inspection.overall_result == "Fail"

    ).count()

    pending = query.filter(

        Inspection.overall_result == "Pending"

    ).count()

    pass_rate = round(

        (passed / total) * 100,

        1

    ) if total else 0

    return {

        "total": total,

        "passed": passed,

        "failed": failed,

        "pending": pending,

        "pass_rate": pass_rate

    }

def export_excel(

        title,

        headers,

        rows,

        filename

):

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = title

    header_fill = PatternFill(

        start_color="1F4E78",

        end_color="1F4E78",

        fill_type="solid"

    )

    header_font = Font(

        bold=True,

        color="FFFFFF"

    )

    # Write headers
    for column, header in enumerate(headers, start=1):

        cell = worksheet.cell(

            row=1,

            column=column

        )

        cell.value = header

        cell.fill = header_fill

        cell.font = header_font

    # Write rows
    current_row = 2

    for row in rows:

        for column, value in enumerate(row, start=1):

            worksheet.cell(

                row=current_row,

                column=column

            ).value = value

        current_row += 1

    # Auto-fit columns
    for column_cells in worksheet.columns:

        max_length = 0

        column_letter = column_cells[0].column_letter

        for cell in column_cells:

            try:

                if cell.value:

                    max_length = max(

                        max_length,

                        len(str(cell.value))

                    )

            except Exception:

                pass

        worksheet.column_dimensions[

            column_letter

        ].width = max_length + 5

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return send_file(

        output,

        as_attachment=True,

        download_name=filename,

        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

def export_pdf(

        title,

        headers,

        rows,

        filename

):

    buffer = BytesIO()

    document = SimpleDocTemplate(

        buffer,

        pagesize=landscape(A4),

        rightMargin=20,

        leftMargin=20,

        topMargin=20,

        bottomMargin=20

    )

    styles = getSampleStyleSheet()

    title_style = styles["Heading1"]

    title_style.alignment = TA_CENTER

    elements = []
    elements.append(

        Paragraph(

            f"<b>{title}</b>",

            title_style

        )

    )

    elements.append(

        Spacer(

            1,

            0.30 * inch

        )

    )
    table_data = [headers]

    for row in rows:

        table_data.append(row)

    table = Table(table_data, repeatRows= 1)
    table.setStyle(

        TableStyle([

            (

                "BACKGROUND",

                (0, 0),

                (-1, 0),

                colors.HexColor("#1E3A8A")

            ),

            (

                "TEXTCOLOR",

                (0, 0),

                (-1, 0),

                colors.white

            ),

            (

                "FONTNAME",

                (0, 0),

                (-1, 0),

                "Helvetica-Bold"

            ),

            (

                "FONTSIZE",

                (0, 0),

                (-1, -1),

                6

            ),

            (

                "GRID",

                (0, 0),

                (-1, -1),

                0.5,

                colors.grey

            ),

            (

                "BACKGROUND",

                (0, 1),

                (-1, -1),

                colors.whitesmoke

            ),

            (

                "ALIGN",

                (0, 0),

                (-1, -1),

                "CENTER"

            ),

            (

                "BOTTOMPADDING",

                (0, 0),

                (-1, 0),

                8

            )

        ])

    )

    elements.append(

        table

    )

    document.build(

        elements

    )

    buffer.seek(0)

    return send_file(

        buffer,

        as_attachment=True,

        download_name=filename,

        mimetype="application/pdf"

    )

#############################################################
# Generic Helpers
#############################################################

def build_rows(

        objects,

        columns

):

    rows = []

    for obj in objects:

        row = []

        for column in columns:

            value = obj

            for attr in column.split("."):

                value = getattr(

                    value,

                    attr,

                    ""

                )

            if hasattr(

                value,

                "strftime"

            ):

                value = value.strftime(

                    "%d-%m-%Y"

                )

            row.append(value)

        rows.append(row)

    return rows

def get_result_chart_data(query):

    passed = query.filter(

        Inspection.overall_result == "Pass"

    ).count()

    failed = query.filter(

        Inspection.overall_result == "Fail"

    ).count()

    pending = query.filter(

        Inspection.overall_result == "Pending"

    ).count()

    return {

        "labels": [

            "Pass",

            "Fail",

            "Pending"

        ],

        "values": [

            passed,

            failed,

            pending

        ]

    }


def get_monthly_trend(company_id):

    results = (
        db.session.query(
            Inspection.inspection_date,
            func.count(Inspection.id)
        )
        .filter(
            Inspection.company_id == company_id
        )
        .group_by(
            Inspection.inspection_date
        )
        .order_by(
            Inspection.inspection_date
        )
        .all()
    )

    labels = [
        date.strftime("%d %b %Y")
        for date, _ in results
    ]

    data = [
        count
        for _, count in results
    ]

    return {
        "labels": labels,
        "data": data
    }
def get_production_line_chart(company_id):

    results = (

        db.session.query(

            ProductionLine.line_name,

            func.count(

                Inspection.id

            )

        )

        .join(

            CableBatch,

            CableBatch.production_line_id == ProductionLine.id

        )

        .join(

            Inspection,

            Inspection.batch_id == CableBatch.id

        )

        .filter(

            Inspection.company_id == company_id

        )

        .group_by(

            ProductionLine.line_name

        )

        .all()

    )

    return {

        "labels":[

            row[0]

            for row in results

        ],

        "values":[

            row[1]

            for row in results

        ]

    }

def get_cable_type_chart(company_id):

    results = (

        db.session.query(

            CableType.name,

            func.count(

                Inspection.id

            )

        )

        .join(

            CableBatch,

            CableBatch.cable_type_id == CableType.id

        )

        .join(

            Inspection,

            Inspection.batch_id == CableBatch.id

        )

        .filter(

            Inspection.company_id == company_id

        )

        .group_by(

            CableType.name

        )

        .all()

    )

    return {

        "labels":[

            row[0]

            for row in results

        ],

        "values":[

            row[1]

            for row in results

        ]

    }

def get_filtered_deviation_query(company_id):

    query = Deviation.query.filter(

        Deviation.company_id == company_id

    )

    if request.args.get("status"):

        query = query.filter(

            Deviation.status == request.args.get("status")

        )

    if request.args.get("severity"):

        query = query.filter(

            Deviation.severity == request.args.get("severity")

        )

    if request.args.get("reported_by"):

        query = query.filter(

            Deviation.reported_by.ilike(

                f"%{request.args.get('reported_by')}%"

            )

        )

    if request.args.get("date_from"):

        query = query.filter(

            Deviation.created_at >= request.args.get("date_from")

        )

    if request.args.get("date_to"):

        query = query.filter(

            Deviation.created_at <= request.args.get("date_to")

        )

    return query


def get_deviation_statistics(query):

    total = query.count()

    open_count = query.filter(

        Deviation.status == "Open"

    ).count()

    progress = query.filter(

        Deviation.status == "In Progress"

    ).count()

    closed = query.filter(

        Deviation.status == "Closed"

    ).count()

    critical = query.filter(

        Deviation.severity == "Critical"

    ).count()

    return {

        "total": total,

        "open": open_count,

        "progress": progress,

        "closed": closed,

        "critical": critical

    }


def get_deviation_chart(query):

    open_count = query.filter(

        Deviation.status == "Open"

    ).count()

    progress = query.filter(

        Deviation.status == "In Progress"

    ).count()

    closed = query.filter(

        Deviation.status == "Closed"

    ).count()

    return {

        "labels": [

            "Open",

            "In Progress",

            "Closed"

        ],

        "values": [

            open_count,

            progress,

            closed

        ]

    }

def get_severity_chart(query):

    minor = query.filter(

        Deviation.severity == "Minor"

    ).count()

    major = query.filter(

        Deviation.severity == "Major"

    ).count()

    critical = query.filter(

        Deviation.severity == "Critical"

    ).count()

    return {

        "labels": [

            "Minor",

            "Major",

            "Critical"

        ],

        "values": [

            minor,

            major,

            critical

        ]

    }

def apply_deviation_sort(query):

    sort = request.args.get(

        "sort",

        "reported_date"

    )

    direction = request.args.get(

        "direction",

        "desc"

    )

    columns = {

        "deviation_number": Deviation.deviation_number,

        "severity": Deviation.severity,

        "status": Deviation.status,

        "reported_by": Deviation.reported_by,

        "reported_date": Deviation.reported_date,

        "closed_date": Deviation.closed_date

    }

    column = columns.get(

        sort,

        Deviation.reported_date

    )

    if direction == "asc":

        query = query.order_by(

            column.asc()

        )

    else:

        query = query.order_by(

            column.desc()

        )

    return query


def get_filtered_capa_query(company_id):

    query = CAPA.query.filter(

        CAPA.company_id == company_id

    )

    if request.args.get("status"):

        query = query.filter(

            CAPA.status == request.args.get("status")

        )

    if request.args.get("effectiveness"):

        query = query.filter(

            CAPA.effectiveness == request.args.get("effectiveness")

        )

    if request.args.get("assigned_to"):

        query = query.filter(

            CAPA.assigned_to.ilike(

                f"%{request.args.get('assigned_to')}%"

            )

        )

    if request.args.get("date_from"):

        query = query.filter(

            CAPA.due_date >= request.args.get("date_from")

        )

    if request.args.get("date_to"):

        query = query.filter(

            CAPA.due_date <= request.args.get("date_to")

        )

    return query

def apply_capa_sort(query):

    sort = request.args.get(

        "sort",

        "due_date"

    )

    direction = request.args.get(

        "direction",

        "desc"

    )

    columns = {

        "id": CAPA.id,

        "assigned_to": CAPA.assigned_to,

        "status": CAPA.status,

        "effectiveness": CAPA.effectiveness,

        "due_date": CAPA.due_date,

        "completion_date": CAPA.completion_date

    }

    column = columns.get(

        sort,

        CAPA.due_date

    )

    if direction == "asc":

        query = query.order_by(

            column.asc()

        )

    else:

        query = query.order_by(

            column.desc()

        )

    return query

def get_capa_status_chart(query):

    return {

        "labels": [

            "Open",

            "In Progress",

            "Completed",

            "Overdue",

            "Closed"

        ],

        "data": [

            query.filter(

                CAPA.status == "Open"

            ).count(),

            query.filter(

                CAPA.status == "In Progress"

            ).count(),

            query.filter(

                CAPA.status == "Completed"

            ).count(),

            query.filter(

                CAPA.status == "Overdue"

            ).count(),

            query.filter(

                CAPA.status == "Closed"

            ).count()

        ]

    }

def get_capa_statistics(query):

    total = query.count()

    open_count = query.filter(
        CAPA.status == "Open"
    ).count()

    progress = query.filter(
        CAPA.status == "In Progress"
    ).count()

    completed = query.filter(
        CAPA.status == "Completed"
    ).count()

    closed = query.filter(
        CAPA.status == "Closed"
    ).count()

    overdue = query.filter(

        CAPA.due_date.isnot(None),
        
        CAPA.due_date <= date.today(),
        
        CAPA.status != "Closed"

    ).count()

    return {
        "total": total,
        "open": open_count,
        "progress": progress,
        "completed": completed,
        "closed": closed,
        "overdue": overdue
    }

def get_capa_effectiveness_chart(query):

    return {

        "labels": [

            "Pending",

            "Effective",

            "Not Effective"

        ],

        "data": [

            query.filter(

                CAPA.effectiveness == "Pending"

            ).count(),

            query.filter(

                CAPA.effectiveness == "Effective"

            ).count(),

            query.filter(

                CAPA.effectiveness == "Not Effective"

            ).count()

        ]

    }


def get_filtered_quality_metric_query(company_id):

    query = QualityMetric.query.filter(

        QualityMetric.company_id == company_id

    )

    if request.args.get("result"):

        query = query.filter(

            QualityMetric.result == request.args.get("result")

        )

    if request.args.get("inspection"):

        query = query.join(

            Inspection

        ).filter(

            Inspection.inspection_number.ilike(

                f"%{request.args.get('inspection')}%"

            )

        )

    if request.args.get("date_from"):

        query = query.join(

            Inspection

        ).filter(

            Inspection.inspection_date >= request.args.get("date_from")

        )

    if request.args.get("date_to"):

        query = query.join(

            Inspection

        ).filter(

            Inspection.inspection_date <= request.args.get("date_to")

        )

    return query

def apply_quality_metric_sort(query):

    sort = request.args.get(

        "sort",

        "result"

    )

    direction = request.args.get(

        "direction",

        "asc"

    )

    columns = {

        "measured_value": QualityMetric.measured_value,

        "result": QualityMetric.result

    }

    column = columns.get(

        sort,

        QualityMetric.result

    )

    if direction == "asc":

        query = query.order_by(

            column.asc()

        )

    else:

        query = query.order_by(

            column.desc()

        )

    return query

def get_quality_metric_statistics(query):

    total = query.count()

    passed = query.filter(

        QualityMetric.result == "Pass"

    ).count()

    failed = query.filter(

        QualityMetric.result == "Fail"

    ).count()

    return {

        "total": total,

        "passed": passed,

        "failed": failed

    }

def get_quality_metric_chart(query):

    return {

        "labels":[

            "Pass",

            "Fail"

        ],

        "values":[

            query.filter(

                QualityMetric.result=="Pass"

            ).count(),

            query.filter(

                QualityMetric.result=="Fail"

            ).count()

        ]

    }

def get_filtered_production_query(company_id):

    query = CableBatch.query.filter(

        CableBatch.company_id == company_id

    )

    if request.args.get("status"):

        query = query.filter(

            CableBatch.status == request.args.get("status")

        )

    if request.args.get("customer"):

        query = query.join(

            Customer

        ).filter(

            Customer.company_name.ilike(

                f"%{request.args.get('customer')}%"

            )

        )

    if request.args.get("production_line"):

        query = query.join(

            ProductionLine

        ).filter(

            ProductionLine.line_name == request.args.get(

                "production_line"

            )

        )

    if request.args.get("date_from"):

        query = query.filter(

            CableBatch.production_date >= request.args.get(

                "date_from"

            )

        )

    if request.args.get("date_to"):

        query = query.filter(

            CableBatch.production_date <= request.args.get(

                "date_to"

            )

        )

    return query

def apply_production_sort(query):

    sort = request.args.get(

        "sort",

        "production_date"

    )

    direction = request.args.get(

        "direction",

        "desc"

    )

    columns = {

        "batch_number": CableBatch.batch_number,

        "drum_number": CableBatch.drum_number,

        "production_date": CableBatch.production_date,

        "status": CableBatch.status,

        "cable_length": CableBatch.cable_length

    }

    column = columns.get(

        sort,

        CableBatch.production_date

    )

    if direction == "asc":

        query = query.order_by(

            column.asc()

        )

    else:

        query = query.order_by(

            column.desc()

        )

    return query

def get_production_statistics(query):

    total = query.count()

    pending = query.filter(

        CableBatch.status == "Pending"

    ).count()

    completed = query.filter(

        CableBatch.status == "Completed"

    ).count()

    rejected = query.filter(

        CableBatch.status == "Rejected"

    ).count()

    total_length = sum(

        batch.cable_length

        for batch in query.all()

    )

    return {

        "total": total,

        "pending": pending,

        "completed": completed,

        "rejected": rejected,

        "total_length": total_length

    }

def get_production_status_chart(query):

    return {

        "labels":[

            "Pending",

            "Completed",

            "Rejected"

        ],

        "values":[

            query.filter(

                CableBatch.status=="Pending"

            ).count(),

            query.filter(

                CableBatch.status=="Completed"

            ).count(),

            query.filter(

                CableBatch.status=="Rejected"

            ).count()

        ]

    }

def get_filtered_customer_query(company_id):

    query = Customer.query.filter(

        Customer.company_id == company_id

    )

    if request.args.get("company_name"):

        query = query.filter(

            Customer.company_name.ilike(

                f"%{request.args.get('company_name')}%"

            )

        )

    if request.args.get("contact_person"):

        query = query.filter(

            Customer.contact_person.ilike(

                f"%{request.args.get('contact_person')}%"

            )

        )

    if request.args.get("email"):

        query = query.filter(

            Customer.email.ilike(

                f"%{request.args.get('email')}%"

            )

        )

    return query

def apply_customer_sort(query):

    sort = request.args.get(

        "sort",

        "company_name"

    )

    direction = request.args.get(

        "direction",

        "asc"

    )

    columns = {

        "company_name": Customer.company_name,

        "contact_person": Customer.contact_person,

        "email": Customer.email,

        "phone": Customer.phone,

        "created_at": Customer.created_at

    }

    column = columns.get(

        sort,

        Customer.company_name

    )

    if direction == "asc":

        query = query.order_by(

            column.asc()

        )

    else:

        query = query.order_by(

            column.desc()

        )

    return query


def get_customer_request_chart(company_id):
    """
    Number of accepted cable requests per customer.
    Rejected batches are excluded.
    """

    results = (
        db.session.query(
            Customer.company_name,
            func.count(CableBatch.id)
        )
        .join(
            CableBatch,
            CableBatch.customer_id == Customer.id
        )
        .filter(
            Customer.company_id == company_id,
            CableBatch.status != "Rejected"
        )
        .group_by(
            Customer.id,
            Customer.company_name
        )
        .order_by(
            func.count(CableBatch.id).desc()
        )
        .all()
    )

    return {
        "labels": [r[0] for r in results],
        "values": [r[1] for r in results]
    }

def get_customer_statistics(query):

    total = query.count()

    with_email = query.filter(

        Customer.email.isnot(None),

        Customer.email != ""

    ).count()

    with_phone = query.filter(

        Customer.phone.isnot(None),

        Customer.phone != ""

    ).count()

    return {

        "total": total,

        "with_email": with_email,

        "with_phone": with_phone

    }

def get_filtered_user_query(company_id):

    query = User.query.filter(

        User.company_id == company_id

    )

    if request.args.get("full_name"):

        query = query.filter(

            User.full_name.ilike(

                f"%{request.args.get('full_name')}%"

            )

        )

    if request.args.get("email"):

        query = query.filter(

            User.email.ilike(

                f"%{request.args.get('email')}%"

            )

        )

    if request.args.get("role"):

        query = query.filter(

            User.role == request.args.get(

                "role"

            )

        )

    if request.args.get("active"):

        active = request.args.get("active") == "true"

        query = query.filter(

            User.is_active == active

        )

    return query

def apply_user_sort(query):

    sort = request.args.get(

        "sort",

        "full_name"

    )

    direction = request.args.get(

        "direction",

        "asc"

    )

    columns = {

        "full_name": User.full_name,

        "username": User.username,

        "email": User.email,

        "role": User.role,

        "created_at": User.created_at

    }

    column = columns.get(

        sort,

        User.full_name

    )

    if direction == "asc":

        query = query.order_by(

            column.asc()

        )

    else:

        query = query.order_by(

            column.desc()

        )

    return query

def get_user_statistics(query):

    total = query.count()

    active = query.filter(

        User.is_active == True

    ).count()

    inactive = query.filter(

        User.is_active == False

    ).count()

    admins = query.filter(

        User.role == "Admin"

    ).count()

    return {

        "total": total,

        "active": active,

        "inactive": inactive,

        "admins": admins

    }

def get_user_status_chart(query):

    return {

        "labels":[

            "Active",

            "Inactive"

        ],

        "values":[

            query.filter(

                User.is_active == True

            ).count(),

            query.filter(

                User.is_active == False

            ).count()

        ]

    }


def get_filtered_audit_query(company_id):

    query = AuditLog.query.filter(

        AuditLog.company_id == company_id

    )

    if request.args.get("module"):

        query = query.filter(

            AuditLog.module.ilike(

                f"%{request.args.get('module')}%"

            )

        )

    if request.args.get("action"):

        query = query.filter(

            AuditLog.action.ilike(

                f"%{request.args.get('action')}%"

            )

        )

    if request.args.get("user"):

        query = query.join(

            User

        ).filter(

            User.full_name.ilike(

                f"%{request.args.get('user')}%"

            )

        )

    if request.args.get("date_from"):

        query = query.filter(

            AuditLog.created_at >= request.args.get(

                "date_from"

            )

        )

    if request.args.get("date_to"):

        query = query.filter(

            AuditLog.created_at <= request.args.get(

                "date_to"

            )

        )

    return query


def apply_audit_sort(query):

    sort = request.args.get(

        "sort",

        "created_at"

    )

    direction = request.args.get(

        "direction",

        "desc"

    )

    columns = {

        "module": AuditLog.module,

        "action": AuditLog.action,

        "created_at": AuditLog.created_at

    }

    column = columns.get(

        sort,

        AuditLog.created_at

    )

    if direction == "asc":

        query = query.order_by(

            column.asc()

        )

    else:

        query = query.order_by(

            column.desc()

        )

    return query

def get_audit_statistics(query):

    total = query.order_by(None).count()

    modules = (
        query
        .order_by(None)
        .with_entities(AuditLog.module)
        .distinct()
        .count()
    )

    actions = (
        query
        .order_by(None)
        .with_entities(AuditLog.action)
        .distinct()
        .count()
    )

    users = (
        query
        .order_by(None)
        .with_entities(AuditLog.user_id)
        .distinct()
        .count()
    )

    return {
        "total": total,
        "modules": modules,
        "actions": actions,
        "users": users
    }
def get_audit_chart(query):

    results = (
        query
        .order_by(None)
        .with_entities(
            AuditLog.module,
            func.count(AuditLog.id)
        )
        .group_by(
            AuditLog.module
        )
        .order_by(
            func.count(AuditLog.id).desc()
        )
        .all()
    )

    return {
        "labels": [
            row[0] for row in results
        ],
        "data": [
            row[1] for row in results
        ]
    }


def export_all_companies_excel():

    """
    Generate a complete system-wide Excel report.

    Sheets:
    1. Company Overview
    2. Users
    3. Production
    4. Inspections
    5. Deviations
    6. CAPA
    """

    workbook = Workbook()

    # =========================================================
    # COMMON STYLING
    # =========================================================

    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    title_font = Font(
        bold=True,
        size=14
    )

    def style_header(worksheet):

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

    def auto_width(worksheet):

        for column_cells in worksheet.columns:

            max_length = 0

            column_letter = column_cells[0].column_letter

            for cell in column_cells:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 4,
                50
            )

    # =========================================================
    # GET ALL DATA
    # =========================================================

    companies = Company.query.order_by(
        Company.company_name
    ).all()

    users = User.query.order_by(
        User.full_name
    ).all()

    batches = CableBatch.query.order_by(
        CableBatch.production_date.desc()
    ).all()

    inspections = Inspection.query.order_by(
        Inspection.inspection_date.desc()
    ).all()

    deviations = Deviation.query.order_by(
        Deviation.reported_date.desc()
    ).all()

    capas = CAPA.query.order_by(
        CAPA.due_date.desc()
    ).all()

    # =========================================================
    # 1. COMPANY OVERVIEW
    # =========================================================

    worksheet = workbook.active

    worksheet.title = "Company Overview"

    worksheet["A1"] = "CableQIMS System-Wide Company Report"

    worksheet["A1"].font = title_font

    worksheet["A3"] = "Total Companies"
    worksheet["B3"] = len(companies)

    worksheet["A4"] = "Active Companies"

    worksheet["B4"] = sum(
        1 for company in companies
        if company.is_active
    )

    worksheet["A5"] = "Inactive Companies"

    worksheet["B5"] = sum(
        1 for company in companies
        if not company.is_active
    )

    worksheet["A6"] = "Total Users"
    worksheet["B6"] = len(users)

    worksheet["A7"] = "Total Production Batches"
    worksheet["B7"] = len(batches)

    worksheet["A8"] = "Total Inspections"
    worksheet["B8"] = len(inspections)

    worksheet["A9"] = "Total Deviations"
    worksheet["B9"] = len(deviations)

    worksheet["A10"] = "Total CAPA"
    worksheet["B10"] = len(capas)

    worksheet["A12"] = "COMPANIES"
    worksheet["A12"].font = title_font

    worksheet.append([
        "Company Name",
        "Company Code",
        "Address",
        "Email",
        "Phone",
        "Status",
        "Created"
    ])

    for company in companies:

        worksheet.append([
            company.company_name,
            company.company_code,
            company.address or "",
            company.email or "",
            company.phone or "",
            "Active"
            if company.is_active
            else "Inactive",
            company.created_at.strftime("%d-%m-%Y")
            if company.created_at
            else ""
        ])

    style_header(worksheet)
    auto_width(worksheet)

    # =========================================================
    # 2. USERS
    # =========================================================

    worksheet = workbook.create_sheet(
        "Users"
    )

    worksheet.append([
        "Company",
        "Company Code",
        "Full Name",
        "Username",
        "Email",
        "Role",
        "Status",
        "Created Date"
    ])

    for user in users:

        worksheet.append([
            user.company.company_name
            if user.company else "",

            user.company.company_code
            if user.company else "",

            user.full_name,
            user.username,
            user.email,
            user.role,

            "Active"
            if user.is_active
            else "Inactive",

            user.created_at.strftime("%d-%m-%Y")
            if user.created_at
            else ""
        ])

    style_header(worksheet)
    auto_width(worksheet)

    # =========================================================
    # 3. PRODUCTION
    # =========================================================

    worksheet = workbook.create_sheet(
        "Production"
    )

    worksheet.append([
        "Company",
        "Company Code",
        "Batch Number",
        "Cable Code",
        "Drum Number",
        "Customer",
        "Cable Name",
        "Production Line",
        "Supervisor",
        "Production Date",
        "Cable Length",
        "Status"
    ])

    for batch in batches:

        worksheet.append([
            batch.company.company_name
            if batch.company else "",

            batch.company.company_code
            if batch.company else "",

            batch.batch_number,
            batch.cable_code,
            batch.drum_number,

            f"{batch.customer.company_name} \n {batch.customer.address}"
            if batch.customer
            else "",

            f"{batch.cable_type.pair_count} x {batch.cable_type.conductor_size} - \n{batch.cable_type.name}"
            if batch.cable_type
            else "",

            batch.production_line.line_name
            if batch.production_line
            else "",

            batch.production_line.supervisor
            if batch.production_line
            else "",

            batch.production_date.strftime(
                "%d-%m-%Y"
            )
            if batch.production_date
            else "",

            batch.cable_length or 0,

            batch.status
        ])

    style_header(worksheet)
    auto_width(worksheet)

    # =========================================================
    # 4. INSPECTIONS
    # =========================================================

    worksheet = workbook.create_sheet(
        "Inspections"
    )

    worksheet.append([
        "Company",
        "Company Code",
        "Inspection Number",
        "Batch Number",
        "Inspector",
        "Inspection Date",
        "Result",
        "Remarks"
    ])

    for inspection in inspections:

        worksheet.append([
            inspection.company.company_name
            if inspection.company
            else "",

            inspection.company.company_code
            if inspection.company
            else "",

            inspection.inspection_number,

            inspection.batch.batch_number
            if inspection.batch
            else "",

            inspection.inspector,

            inspection.inspection_date.strftime(
                "%d-%m-%Y"
            )
            if inspection.inspection_date
            else "",

            inspection.overall_result,

            inspection.remarks or ""
        ])

    style_header(worksheet)
    auto_width(worksheet)

    # =========================================================
    # 5. DEVIATIONS
    # =========================================================

    worksheet = workbook.create_sheet(
        "Deviations"
    )

    worksheet.append([
        "Company",
        "Company Code",
        "Deviation Number",
        "Inspection",
        "Description",
        "Severity",
        "Status",
        "Reported By",
        "Reported Date",
        "Closed Date"
    ])

    for deviation in deviations:

        worksheet.append([
            deviation.company.company_name
            if deviation.company
            else "",

            deviation.company.company_code
            if deviation.company
            else "",

            deviation.deviation_number,

            deviation.inspection.inspection_number
            if deviation.inspection
            else "",

            deviation.description,

            deviation.severity,
            deviation.status,
            deviation.reported_by,

            deviation.reported_date.strftime(
                "%d-%m-%Y"
            )
            if deviation.reported_date
            else "",

            deviation.closed_date.strftime(
                "%d-%m-%Y"
            )
            if deviation.closed_date
            else ""
        ])

    style_header(worksheet)
    auto_width(worksheet)

    # =========================================================
    # 6. CAPA
    # =========================================================

    worksheet = workbook.create_sheet(
        "CAPA"
    )

    worksheet.append([
        "Company",
        "Company Code",
        "CAPA ID",
        "Deviation",
        "Corrective Action",
        "Preventive Action",
        "Assigned To",
        "Due Date",
        "Completion Date",
        "Effectiveness",
        "Status"
    ])

    for capa in capas:

        worksheet.append([
            capa.company.company_name
            if capa.company
            else "",

            capa.company.company_code
            if capa.company
            else "",

            capa.id,

            capa.deviation.deviation_number
            if capa.deviation
            else "",

            capa.corrective_action,
            capa.preventive_action,
            capa.assigned_to,

            capa.due_date.strftime(
                "%d-%m-%Y"
            )
            if capa.due_date
            else "",

            capa.completion_date.strftime(
                "%d-%m-%Y"
            )
            if capa.completion_date
            else "",

            capa.effectiveness,
            capa.status
        ])

    style_header(worksheet)
    auto_width(worksheet)

    # =========================================================
    # CREATE FILE
    # =========================================================

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="CableQIMS_All_Companies_Report.xlsx",
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )


def export_all_companies_pdf():

    # Get every company
    companies = Company.query.order_by(
        Company.company_name
    ).all()

    # Get all records
    users = User.query.order_by(
        User.full_name
    ).all()

    batches = CableBatch.query.order_by(
        CableBatch.production_date.desc()
    ).all()

    inspections = Inspection.query.order_by(
        Inspection.inspection_date.desc()
    ).all()

    deviations = Deviation.query.order_by(
        Deviation.reported_date.desc()
    ).all()

    capas = CAPA.query.order_by(
        CAPA.due_date.desc()
    ).all()

    buffer = BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=25,
        leftMargin=25,
        topMargin=25,
        bottomMargin=25
    )

    styles = getSampleStyleSheet()

    title_style = styles["Title"]
    title_style.alignment = TA_CENTER

    heading_style = styles["Heading2"]

    elements = []

    # =========================================================
    # TITLE
    # =========================================================

    elements.append(
        Paragraph(
            "CableQIMS System-Wide Company Report",
            title_style
        )
    )

    elements.append(
        Spacer(1, 0.2 * inch)
    )

    elements.append(
        Paragraph(
            "System Administrator Report — All Companies",
            heading_style
        )
    )

    elements.append(
        Spacer(1, 0.2 * inch)
    )

    # =========================================================
    # SUMMARY
    # =========================================================

    total_companies = len(companies)

    active_companies = sum(
        1 for company in companies
        if company.is_active
    )

    total_users = len(users)

    active_users = sum(
        1 for user in users
        if user.is_active
    )

    total_batches = len(batches)

    completed_batches = sum(
        1 for batch in batches
        if batch.status == "Completed"
    )

    total_cable_length = sum(
        (batch.cable_length or 0)
        for batch in batches
    )

    total_inspections = len(inspections)

    passed_inspections = sum(
        1 for inspection in inspections
        if inspection.overall_result == "Pass"
    )

    failed_inspections = sum(
        1 for inspection in inspections
        if inspection.overall_result == "Fail"
    )

    pending_inspections = (
        total_inspections
        - passed_inspections
        - failed_inspections
    )

    inspection_pass_rate = (
        round(
            (passed_inspections / total_inspections) * 100,
            2
        )
        if total_inspections
        else 0
    )

    total_deviations = len(deviations)

    open_deviations = sum(
        1 for deviation in deviations
        if deviation.status == "Open"
    )

    total_capa = len(capas)

    open_capa = sum(
        1 for capa in capas
        if capa.status == "Open"
    )

    summary_data = [
        ["Metric", "Value"],

        ["Companies", total_companies],

        ["Active Companies", active_companies],

        ["Users", total_users],

        ["Active Users", active_users],

        ["Production Batches", total_batches],

        ["Completed Batches", completed_batches],

        [
            "Cable Produced",
            f"{total_cable_length:,.2f} metres"
        ],

        ["Inspections", total_inspections],

        ["Passed Inspections", passed_inspections],

        ["Failed Inspections", failed_inspections],

        ["Pending Inspections", pending_inspections],

        [
            "Inspection Pass Rate",
            f"{inspection_pass_rate}%"
        ],

        ["Deviations", total_deviations],

        ["Open Deviations", open_deviations],

        ["CAPA", total_capa],

        ["Open CAPA", open_capa],
    ]

    summary_table = Table(
        summary_data,
        colWidths=[4 * inch, 2.5 * inch]
    )

    summary_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#1F4E78")
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),
            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey
            ),
            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                8
            ),
            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "TOP"
            )
        ])
    )

    elements.append(summary_table)

    elements.append(
        Spacer(1, 0.3 * inch)
    )

    # =========================================================
    # GENERIC SECTION
    # =========================================================

    def add_section(title, headers, rows):

        elements.append(
            Paragraph(
                title,
                heading_style
            )
        )

        elements.append(
            Spacer(1, 0.08 * inch)
        )

        if not rows:

            elements.append(
                Paragraph(
                    "No records found.",
                    styles["Normal"]
                )
            )

            elements.append(
                Spacer(1, 0.2 * inch)
            )

            return

        table = Table(
            [headers] + rows,
            repeatRows=1
        )

        table.setStyle(
            TableStyle([
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#1F4E78")
                ),
                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white
                ),
                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "Helvetica-Bold"
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.4,
                    colors.grey
                ),
                (
                    "FONTSIZE",
                    (0, 0),
                    (-1, -1),
                    6.5
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "TOP"
                )
            ])
        )

        elements.append(table)

        elements.append(
            Spacer(1, 0.3 * inch)
        )

    # =========================================================
    # COMPANIES
    # =========================================================

    company_rows = []

    for company in companies:

        company_rows.append([
            company.company_name,
            company.company_code,
            company.address or "",
            company.email or "",
            company.phone or "",
            "Active"
            if company.is_active
            else "Inactive",
            company.created_at.strftime("%d-%m-%Y")
            if company.created_at
            else ""
        ])

    add_section(
        "Companies",
        [
            "Company",
            "Code",
            "Address",
            "Email",
            "Phone",
            "Status",
            "Created"
        ],
        company_rows
    )

    # =========================================================
    # USERS
    # =========================================================

    user_rows = []

    for user in users:

        user_rows.append([
            user.company.company_name
            if user.company
            else "",
            user.full_name,
            user.username,
            user.email,
            user.role,
            "Active"
            if user.is_active
            else "Inactive"
        ])

    add_section(
        "Users",
        [
            "Company",
            "Full Name",
            "Username",
            "Email",
            "Role",
            "Status"
        ],
        user_rows
    )

    # =========================================================
    # PRODUCTION
    # =========================================================

    production_rows = []

    for batch in batches:

        production_rows.append([
            batch.company.company_name
            if batch.company
            else "",
            batch.batch_number,
            batch.cable_code,
            batch.drum_number,
            batch.customer.company_name
            if batch.customer
            else "",
            f"{batch.cable_type.pair_count} x {batch.cable_type.conductor_size} - \n{batch.cable_type.name}"
            if batch.cable_type
            else "",
            batch.production_line.line_name
            if batch.production_line
            else "",
            batch.production_line.supervisor
            if batch.production_line
            else "",
            batch.production_date.strftime("%d-%m-%Y")
            if batch.production_date
            else "",
            f"{batch.cable_length or 0:,.2f}",
            batch.status
        ])

    add_section(
        "Production",
        [
            "Company",
            "Batch",
            "Cable Code",
            "Drum",
            "Customer",
            "Cable Name",
            "Production Line",
            "Supervisor",
            "Date",
            "Length",
            "Status"
        ],
        production_rows
    )

    # =========================================================
    # INSPECTIONS
    # =========================================================

    inspection_rows = []

    for inspection in inspections:

        inspection_rows.append([
            inspection.company.company_name
            if inspection.company
            else "",
            inspection.inspection_number,
            inspection.batch.batch_number
            if inspection.batch
            else "",
            inspection.inspector,
            inspection.inspection_date.strftime("%d-%m-%Y")
            if inspection.inspection_date
            else "",
            inspection.overall_result
        ])

    add_section(
        "Inspections",
        [
            "Company",
            "Inspection",
            "Batch",
            "Inspector",
            "Date",
            "Result"
        ],
        inspection_rows
    )

    # =========================================================
    # DEVIATIONS
    # =========================================================

    deviation_rows = []

    for deviation in deviations:

        deviation_rows.append([
            deviation.company.company_name
            if deviation.company
            else "",
            deviation.deviation_number,
            deviation.inspection.inspection_number
            if deviation.inspection
            else "",
            deviation.description,
            deviation.severity,
            deviation.status,
            deviation.reported_by,
            deviation.reported_date.strftime("%d-%m-%Y")
            if deviation.reported_date
            else ""
        ])

    add_section(
        "Deviations",
        [
            "Company",
            "Deviation",
            "Inspection",
            "Description",
            "Severity",
            "Status",
            "Reported By",
            "Date"
        ],
        deviation_rows
    )

    # =========================================================
    # CAPA
    # =========================================================

    capa_rows = []

    for capa in capas:

        capa_rows.append([
            capa.company.company_name
            if capa.company
            else "",
            capa.deviation.deviation_number
            if capa.deviation
            else "",
            capa.assigned_to,
            capa.due_date.strftime("%d-%m-%Y")
            if capa.due_date
            else "",
            capa.completion_date.strftime("%d-%m-%Y")
            if capa.completion_date
            else "",
            capa.effectiveness,
            capa.status
        ])

    add_section(
        "CAPA",
        [
            "Company",
            "Deviation",
            "Assigned To",
            "Due Date",
            "Completion",
            "Effectiveness",
            "Status"
        ],
        capa_rows
    )

    # =========================================================
    # BUILD PDF
    # =========================================================

    document.build(elements)

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="CableQIMS_All_Companies_Report.pdf",
        mimetype="application/pdf"
    )






